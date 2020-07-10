# импорт необходимых библиотек библиотек
from selenium import webdriver
from selenium.webdriver.common.proxy import Proxy, ProxyType
from openpyxl import load_workbook
import time


def retry(func):
    def wrapper():
        try:
            func()

        except Exception as e:
            print(e)
            time.sleep(60)
            func()
    return wrapper


# Главная функция обернутая в ретрай декоратор
@retry
def main():

    # Настройка параметров openpyxl для выгрузки в csv и загрузка Excel шаблона для данных
    wb = load_workbook('excel_template\\ExcelTemplateForDownloadingNewsFromZakonkz.xlsx')
    ws = wb.active

    # !!! Блок для использования прокси при парсинге !!! #

    # IP взято с сайта http://spys.one/
    # ip = "180.183.10.210:8213"

    # prox = Proxy()
    # prox.proxy_type = ProxyType.MANUAL
    # prox.http_proxy = ip
    # prox.socks_proxy = ip
    # prox.ssl_proxy = ip

    # capabilities = webdriver.DesiredCapabilities.CHROME
    # prox.add_to_capabilities(capabilities)

    # browser = webdriver.Chrome("chromedriver_folder\\chromedriver.exe", desired_capabilities=capabilities)
    # browser.maximize_window()
    # !!!                                           !!! #


    browser = webdriver.Chrome("chromedriver_folder\\chromedriver.exe")
    browser.maximize_window()

    # Получение всего html наполнения страницы
    news_list_link = "https://www.zakon.kz/news"
    browser.get(news_list_link)

    # # # Получение  списка ссылок на текущие статьи и даты их публикации # # #
    # Разбиения класса содержащего в себе все ссылки на отдельные объекты
    all_current_news_objects = browser.find_element_by_xpath('//*[@id="dle-content"]').find_elements_by_class_name('cat_news_item')

    # Вычленение ссылки из объектов списка all_current_news_objects, исключая первый объектов, так как в нём содержится дата публикаций ниже, а не ссылка на статьи
    # Для каждого элемента проводится следующая обработка: переход в тег <a> -> получение ссылки на статью
    # Так как на странице zakon.kz/news в определенный момент могут присутствовать новости сразу за два дня,
    # реализована проверка на проверку даты новостей, чтобы спарсить в итоге только новости за текущий день
    all_current_news_links = []
    for i in all_current_news_objects[1:]:
        if i.text[4] == '-':
            break

        all_current_news_links.append(i.find_element_by_tag_name('a').get_attribute('href'))

    # Получение даты публикаций
    current_news_date = all_current_news_objects[0].text

    # Флаг для обновления строки записи в Excel
    excel_row = 2

    # Отдельный парсинг статей по ссылкам
    for news_link in all_current_news_links:
        # # # Парсинг статьи # # #
        # Загружаем страницу со статьей
        browser.get(news_link  )

        # На сайте zakon.kz статьи могут быть офрмлены в двух вариантах: старом и новом
        # Блок для обработки данных с последним дизайном сайта
        try:
            # Получение заголовка статьи
            article_header = browser.find_element_by_xpath('//*[@id="dle-content"]/div[1]/h1').text

            # Получение полного текста статьи
            article_text_content = browser.find_element_by_xpath('//*[@id="initial_news_story"]').text

            # # # Скролл страницы до конца страницы # # #
            # Программа бесконечно скролит страницу вниз покане удостовериться что блок с комментариями прогрузился
            i = 0
            while True:
                if i % 500 == 0:
                    # Проверка того, что блок с комментариями загрузился, прерывания скроинга после этого
                    try:
                        article_comments_count = browser.find_element_by_xpath('//*[@id="zkn_comments"]/div/div[1]/div[1]/span').text
                        print('breaked')
                        break

                    except:
                        pass

                i += 10

                # Напрямую обращение к скрипту JS для прокрутки на определнное количество единиц
                browser.execute_script("window.scrollTo(0, %s);" % i)

            # Получение количества комментариев к статье
            article_comments_count = browser.find_element_by_xpath('//*[@id="zkn_comments"]/div/div[1]/div[1]/span').text

            # Получение коментариев
            # Если комментарии отсутсвуют, в таблицу Excel будет занесено значение NULL
            if article_comments_count != '0':

                # Выделение текста комментариев из блока с комментариями
                # В дальнейшем, при необходимости эту строку с комментариями возможно будет распарсить
                all_article_comments = browser.find_element_by_css_selector('.zknc.zknc-posts').text

            else:
                all_article_comments = 'NULL'

            # Запись данных в таблицу
            ws.cell(row=excel_row, column=1).value = article_header
            ws.cell(row=excel_row, column=2).value = current_news_date
            ws.cell(row=excel_row, column=3).value = article_text_content
            ws.cell(row=excel_row, column=4).value = article_comments_count
            ws.cell(row=excel_row, column=5).value = all_article_comments

            excel_row += 1

        # Исключение для парсинга статьи на старой версии сайта (они всё ещё присутсвуют)
        except Exception as e:
            print(e)

            try:
                # Получение заголовка статьи
                article_header = browser.find_element_by_xpath('//*[@id="dle-content"]/section/div[2]/h1').text

                # Получение полного текста статьи
                article_text_content = browser.find_element_by_xpath('//*[@id="dle-content"]/section/div[2]/div[5]').text

                # # # Скролл страницы до конца страницы # # #
                # Программа бесконечно скролит страницу вниз покане удостовериться что блок с комментариями прогрузился
                i = 0
                while True:
                    if i % 500 == 0:
                        # Проверка того, что блок с комментариями загрузился, прерывания скроинга после этого
                        try:
                            article_comments_count = browser.find_element_by_xpath('//*[@id="zkn_comments"]/div/div[1]/div[1]/span').text
                            print('breaked')
                            break

                        except:
                            pass

                    i += 10

                    # Напрямую обращение к скрипту JS для прокрутки на определнное количество единиц
                    browser.execute_script("window.scrollTo(0, %s);" % i)

                # Получение количества комментариев к статье
                article_comments_count = browser.find_element_by_xpath('//*[@id="zkn_comments"]/div/div[1]/div[1]/span').text

                # Получение коментариев
                # Если комментарии отсутсвуют, в таблицу Excel будет занесено значение NULL
                if article_comments_count != '0':

                    # Выделение текста комментариев из блока с комментариями
                    # В дальнейшем, при необходимости эту строку с комментариями возможно будет распарсить
                    all_article_comments = browser.find_element_by_css_selector('.zknc.zknc-posts').text

                else:
                    all_article_comments = 'NULL'

                # Запись данных в таблицу
                ws.cell(row=excel_row, column=1).value = article_header
                ws.cell(row=excel_row, column=2).value = current_news_date
                ws.cell(row=excel_row, column=3).value = article_text_content
                ws.cell(row=excel_row, column=4).value = article_comments_count
                ws.cell(row=excel_row, column=5).value = all_article_comments

                excel_row += 1
            except Exception as e:
                print(e)

    # Формирование названия и сохранение таблицы в папке проекта
    wb_name = 'NewsFromZakonkz.csv'
    wb.save(wb_name)


# Конструкция для удобного импортирования данного файла в дальнешем
if __name__ == '__main__':
    main()
