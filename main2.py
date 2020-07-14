import requests
from bs4 import BeautifulSoup
import json
import lxml
from openpyxl import load_workbook
import time


def retry(func):
    def wrapper():
        try:
            func()

        except Exception as e:
            print(e)
            time.sleep(60)
            func()
    return wrapper


# Главная функция обернутая в ретрай декоратор
@retry
def main():
    # Использование прокси  при парсинге
    proxy = {
        'http': '118.174.196.112:36314',
        'https': '118.174.196.112:36314'
    }
    # Настройка параметров openpyxl для выгрузки в csv и загрузка Excel шаблона для данных
    wb = load_workbook('excel_template\\ExcelTemplateForDownloadingNewsFromZakonkz.xlsx')
    ws = wb.active

    # Получение всего html наполнения страницы и парсинг его с помощью bs4
    main_list_url = 'https://www.zakon.kz/news'
    main_list_r = requests.get(main_list_url, proxies=proxy)

    main_list_soup = BeautifulSoup(main_list_r.text, 'lxml')
    news_obj_from_main_list = main_list_soup.find('div', 'notmain white_block').find_all('div', 'cat_news_item')

    # Формирование списка из ссылок на текущие статьи
    # Если на странице будут новости за два дня, цикл возьмёт все новости за последнюю дату,
    # и остановить наткнувшись на новую дату
    news_links_from_main_list = []

    for i in news_obj_from_main_list[1:]:
        try:

            news_links_from_main_list.append(' http://zakon.kz'+i.find('a').get('href'))

        except Exception as e:
            print(e)
            # '2020-07-14' насколько значение соотвествует дате, проверяется по тире
            if i.text[4] == '-' and i.text[7] == '-':
                break

    # print(news_links_from_main_list)

    # Получение даты публикаций
    current_news_date = news_obj_from_main_list[0].text

    excel_row = 2

    # Поочередный парсинг статей
    for article_link in news_links_from_main_list:
        try:
            article_html_r = requests.get(article_link, proxies=proxy)
            article_soup = BeautifulSoup(article_html_r.text, 'lxml')

            # Получение заголовка статьи
            article_header = article_soup.find('div', 'fullnews white_block').find('h1').text
            # print(article_header)

            # Получение полного текста статьи
            article_text_content = article_soup.find('div', class_='full_text').text
            # print(article_text_content)

            # print(current_news_date)

            # Запрос нв получение информации о комментариях
            article_zakonnewsid = 'zakonnewsid' + article_link.split('/')[-1].split('-')[0]

            r = requests.post('https://zcomments.net/service/init/1',      data={'page_title': article_header,
                                                                           'page_url': article_link,
                                                                           'block_code': article_zakonnewsid,
                                                                           'lang': 'ru'})

            # Получение комментариев
            article_comments = r.json()['comments']['items']
            article_comments_count = len(article_comments)

            # print(''.join(json.dumps(article_comments)))
            # print(article_comments_count)

            # time.sleep(500)
            # Запись данных в Excel таблицу

            # Запись данных в таблицу
            ws.cell(row=excel_row, column=1).value = article_header
            ws.cell(row=excel_row, column=2).value = current_news_date
            ws.cell(row=excel_row, column=3).value = article_text_content
            ws.cell(row=excel_row, column=4).value = article_comments_count
            ws.cell(row=excel_row, column=5).value = (''.join(json.dumps(article_comments)))

            excel_row += 1

        except Exception as e:
            print(e)
            pass

    # Формирование названия и сохранение таблицы в папке проекта
    wb_name = 'NewsFromZakonkz(type2).csv'
    wb.save(wb_name)


# Конструкция для удобного импортирования данного файла в дальнешем
if __name__ == '__main__':
    main()




